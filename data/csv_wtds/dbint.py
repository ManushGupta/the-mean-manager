
import os
import csv
from openpyxl import load_workbook,Workbook

total_rows = 0
wb1 = Workbook()
ws1=wb1.active

ws1.title= "Integrated Cramer Viznet M6"

ws1['A1'] = 'tiger_id'
ws1['B1'] = 'viznetM6id'
ws1['C1'] = 'gam_id'
ws1['H1'] = 'customer_name'
ws1['I1'] = 'provisioner_name'
ws1['J1'] = 'service_type'
ws1['k1'] = 'inbasket_date'

wb2 = load_workbook('gam.xlsx')
print wb2.get_sheet_names()
ws2 = wb2.get_sheet_by_name("Ovl Funnel Data")
c= ws2.rows
print ws2.cell(row = 1, column = 1).value

for row in c:
	total_rows +=1

dirs = os.listdir(os.getcwd())
print dirs
count_2 = 1

for f in dirs:
	if(f=="Cramer.csv"):
		with open(f,'rb') as fi:
			mycsv = csv.reader(fi)
			count_2-= 1
			for rowcsv in mycsv:
				#print count_2
				#print rowcsv[4]
				
				count_2 +=1
			
				#print rowcsv[4]
				count = 0
				if(rowcsv[4]!='tiger_id'):

					for row in c:
						if(count <= 5):
						
							#print(row[81].value)
							if(row[81].value == rowcsv[4]):
								#print 'no'
								if(row[83].value == rowcsv[4]):
									ws1.cell(row = count_2, column = 1).set_explicit_value(value=rowcsv[4],data_type = 'n') 
									ws1.cell(row = count_2, column = 2).value = row[82].value
									ws1.cell(row = count_2, column = 3).set_explicit_value(value=rowcsv[4],data_type = 'n') 
									ws1.cell(row = count_2,column = 8).value = rowcsv[5]
									ws1.cell(row = count_2, column = 9).value = rowcsv[8]
									ws1.cell(row = count_2, column = 10).value = rowcsv[6]
									ws1.cell(row=count_2,column = 11).value = rowcsv[9]

							else:
								#print 'yes'
								#print count
								if(count==5):
									#print 'yes'
									ws1.cell(row = count_2, column = 1).set_explicit_value(value=rowcsv[4],data_type = 'n') 
									#ws1.cell(row = count_2, column = 2).value = None
									ws1.cell(row = count_2, column = 3).set_explicit_value(value=rowcsv[4],data_type = 'n') 
									ws1.cell(row = count_2,column = 8).value = rowcsv[5]
									ws1.cell(row = count_2, column = 9).value = rowcsv[8]
									ws1.cell(row = count_2, column = 10).value = rowcsv[6]
									ws1.cell(row=count_2,column = 11).value = rowcsv[9]
									




						count +=1
	elif(f=="Viznet.csv"):
		with open(f,'rb') as fi:
			mycsv = csv.reader(fi)
			count_2-=1
			for rowcsv in mycsv:
				#print count_2
				#print rowcsv[0]
				count_2 +=1
				#print count_2
			
				#print rowcsv[0]
				count = 0
				if(rowcsv[0]!="viznet_id"):

					for row in c:
						#print count
						if(count <= 13):
						
							#print(row[81].value)
							if(row[82].value == rowcsv[0]):
								print 'no'
								if(row[83].value == rowcsv[0]):
									ws1.cell(row = count_2, column = 1).value = row[81].value
									ws1.cell(row = count_2, column = 2).value = row[82].value
									ws1.cell(row = count_2, column = 3).value = row[82].value
									ws1.cell(row = count_2,column = 8).value = rowcsv[6]
									ws1.cell(row=count_2,column = 10).value = rowcsv[15]
									ws1.cell(row = count_2,column = 11).value = rowcsv[1]
							else:
								#print 'yes'
								#print count
								if(count==13):
									#print 'yes'
									#ws1.cell(row = count_2, column = 1).value = None
									ws1.cell(row = count_2, column = 2).value = rowcsv[0]
									ws1.cell(row = count_2, column = 3).value = rowcsv[0]
									ws1.cell(row = count_2,column = 8).value = rowcsv[6]
									ws1.cell(row = count_2,column = 10).value = rowcsv[15]
									ws1.cell(row = count_2,column = 11).value = rowcsv[1]




						count +=1
	elif(f=="M6.csv"):
		with open(f,'rb') as fi:
			mycsv = csv.reader(fi)
			count_2-=1
			for rowcsv in mycsv:
				#print count_2
				#print rowcsv[0]
				count_2 +=1
				#print count_2
			
				#print rowcsv[0]
				count = 0
				if(rowcsv[12]!="m6_copf_id"):

					for row in c:
						#print count
						if(count <= 13):
						
							#print(row[81].value)
							if(row[82].value == rowcsv[6][2:]+'_'+rowcsv[12]):
								print 'no'
								#print rowcsv[6][2:]
								if(row[83].value == rowcsv[12]):
									ws1.cell(row = count_2, column = 1).value = row[81].value
									ws1.cell(row = count_2, column = 2).value = row[82].value
									ws1.cell(row = count_2, column = 3).value = row[82].value
									ws1.cell(row = count_2,column = 8).value = rowcsv[0]
									ws1.cell(row = count_2,column = 10).value = rowcsv[1]
									ws1.cell(row = count_2,column = 11).value = rowcsv[27]
							else:
								#print 'yes'
								#print count
								if(count==13):
									#print 'yes'
									#print rowcsv[6][2:] + '_' + rowcsv[12]
									#ws1.cell(row = count_2, column = 1).value = None
									ws1.cell(row = count_2, column = 2).value = rowcsv[12]
									ws1.cell(row = count_2, column = 3).value = rowcsv[12]
									ws1.cell(row = count_2,column = 8).value = rowcsv[0]
									ws1.cell(row = count_2,column = 10).value = rowcsv[1]
									ws1.cell(row = count_2,column = 11).value = rowcsv[27]




						count +=1

	elif(f=="M6_OPP.csv"):
		with open(f,'rb') as fi:
			mycsv = csv.reader(fi)
			count_2-=1
			for rowcsv in mycsv:
				#print count_2
				#print rowcsv[0]
				count_2 +=1
				#print count_2
			
				#print rowcsv[0]
				count = 0
				if(rowcsv[12]!="m6_copf_id"):

					for row in c:
						#print count
						if(count <= 13):
						
							#print(row[81].value)
							if(row[82].value == rowcsv[9][2:]+'_'+rowcsv[5]):
								print 'no'
								#print rowcsv[6][2:]
								if(row[83].value == rowcsv[12]):
									ws1.cell(row = count_2, column = 1).value = row[81].value
									ws1.cell(row = count_2, column = 2).value = row[82].value
									ws1.cell(row = count_2, column = 3).value = row[82].value
									ws1.cell(row = count_2,column = 8).value = rowcsv[0]
									ws1.cell(row = count_2,column = 10).value = rowcsv[4]
									ws1.cell(row = count_2,column = 11).value = rowcsv[43]
							else:
								#print 'yes'
								#print count
								if(count==13):
									#print 'yes'
									#print rowcsv[6][2:] + '_' + rowcsv[12]
									#ws1.cell(row = count_2, column = 1).value = None
									ws1.cell(row = count_2, column = 2).value = rowcsv[12]
									ws1.cell(row = count_2, column = 3).value = rowcsv[12]
									ws1.cell(row = count_2,column = 8).value = rowcsv[0]
									ws1.cell(row = count_2,column = 10).value = rowcsv[4]
									ws1.cell(row = count_2,column = 11).value = rowcsv[43]




						count +=1


wb1.save('compiled_orders.xlsx')