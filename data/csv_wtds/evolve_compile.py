
import os
import csv
from openpyxl import load_workbook,Workbook

count = 0


wb1 = load_workbook('compiled_orders.xlsx')
ws1 = wb1.get_sheet_by_name("Integrated Cramer Viznet M6")

wb2 = load_workbook('evolve.xlsx')
#print wb2.get_sheet_names()
ws2 = wb2.get_sheet_by_name("New Orders")
c1 = ws1.rows
c2= ws2.rows

ws1['D1'] = "mrc_converted"
ws1['E1'] = "nrc_converted"
ws1['F1'] = "program_manager"
ws1['G1'] = "evolve_status"
for row1 in c1:
	#print row1[2].value
	count+= 1
	if(row1[0].value!='tiger_id'):
		for row2 in c2:
			if(row2[1].value!='Opportunity Name'):
				#print row1[2].value
				#print row2[9].value

				
				
				if(row2[9].value == row1[2].value):
					#print 'yes'
					ws1.cell(row = count, column = 4).value = row2[37].value
					ws1.cell(row = count, column = 5).value = row2[39].value
					ws1.cell(row = count, column = 6).value = row2[10].value
					ws1.cell(row = count, column = 7).value = row2[46].value



wb1.save("compiled_orders_evolve.xlsx")	
