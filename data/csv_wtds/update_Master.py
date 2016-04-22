import os
import csv
import numpy as np
import xlrd
from openpyxl import load_workbook,Workbook
dirs = os.listdir(os.getcwd())

wb2 = load_workbook('evolve.xlsx')
#print wb2.get_sheet_names()
ws2 = wb2.get_sheet_by_name("New Orders")
#c1 = ws1.rows
c= ws2.rows

wb2.save("evolve.xls")
count = 0


wb = Workbook()
ws = wb.worksheets[0]
for f in dirs:
	if(f=="dbExport.csv"):
		with open(f,'rb') as fi:
			mycsv = csv.reader(fi)
			count_row = 0
			for row in mycsv:
				count_row +=1
				count_column = 0
				for column in row:
					count_column +=1
					print column
					try:
						ws.cell(row = count_row, column = count_column).value = int(column)
					except ValueError:
						ws.cell(row = count_row, column = count_column).value = (column)

			    		
wb.save("dbExport.xlsx")

#wb1 = xlrd.open_workbook("dbExport.xls")
#ws1 = wb1.sheet_by_index(0)
#c1 = ws1.row_values

wb1 = load_workbook('dbExport.xlsx')
#print wb2.get_sheet_names()
ws1 = wb1.get_sheet_by_name("Sheet")
#c1 = ws1.rows
c1= ws1.rows

wb3 = open_workbook('Cramer_completed.xls')
ws3 = wb3.sheet_by_index(0)
c3 = ws3.row_values

wb4 = open_workbook('M6_completed.xls')
ws4 = wb4.sheet_by_index(0)
c4 = ws4.row_values
for row1 in c1:
	#print row1[2].value
	for row in c:
		#print row[9].value
		if(row1[2].value==row[9].value):
			#print 'yes'
			#print row1[2].value
			if(row1[6].value!=row[46].value):
				print 'yes'
				print row1[6].value
				print row[46].value
				#print row[46].value
				if(row[46].value=="Commissioned" and row1[6].value =="Work in Progress"):
					#print 'yes'
					row1[6].value= "Newly Commissioned"
				
				elif(row[46].value=="Commissioned" and row1[6] == "Newly Commissioned"):
					row1[6].value= "Commissioned"
	for row3 in c3:
		if(row1[2].value == row3[4].value):
			if

	for row4 in range(ws4.nrows):
		if(row1[2].value == c4(row4)[9][2:] + '_' + c4(row4)[5]):
			row1[24].value = "Completed"

wb1.save("updated_master.xlsx")
